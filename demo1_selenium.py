import time
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

file = '/home/kartikey/PycharmProjects/Git_Jenkins_Selenium_Python/Demo_Pipeline_Automation/all_paths_login_logout.xlsx'
sheet = 'Sheet1'
workbook = openpyxl.load_workbook(file)
sheet = workbook[sheet]

driver = webdriver.Chrome(executable_path=ChromeDriverManager().install())
# driver.get("https://practicetestautomation.com/practice-test-login/")
driver.get(sheet.cell(row=2, column=1).value)
driver.maximize_window()
# driver.find_element_by_id('username').send_keys('student')
driver.find_element_by_id(sheet.cell(row=2, column=2).value).send_keys(sheet.cell(row=2, column=3).value)
# driver.find_element_by_id('password').send_keys('Password123')
driver.find_element_by_id(sheet.cell(row=2, column=4).value).send_keys(sheet.cell(row=2, column=5).value)
# driver.find_element_by_id('submit').click()
driver.find_element_by_id(sheet.cell(row=2, column=6).value).click()
time.sleep(2)
# driver.find_element_by_xpath('//*[@id="loop-container"]/div/article/div[2]/div/div/div/a').click()
driver.find_element_by_xpath(sheet.cell(row=2, column=7).value).click()
driver.close()
print("HelloWorld")
print("Hi Kartikey_Tester changed World")
